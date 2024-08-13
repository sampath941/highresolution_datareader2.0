from .db_convert import combine_and_convert_df
import os
import logging
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
import datetime
from flask import Flask, request, send_file, redirect, url_for, flash
from config import Config
import shutil


def save_as_csv(db_path, eventcodes_path, base_filename):
    final_dfs = combine_and_convert_df(db_path, eventcodes_path)
    
    if not final_dfs:
        print("No data fetched to save.")
        return
    # shutil.rmtree(Config.TEMPORARY_FILES_DIR)
    save_dir = Config.TEMPORARY_FILES_DIR # Temporary directory to save the files
    os.makedirs(save_dir, exist_ok=True)
    
    file_paths = []
    
    for table_name, df in final_dfs.items():
        filename = f"{base_filename}_{table_name}.csv"
        file_path = os.path.join(save_dir, filename)
        df.to_csv(file_path, index=False)
        file_paths.append(file_path)
        print(f"Data from table {table_name} exported to CSV at {file_path}")
        logging.info(f"Data from table {table_name} exported to CSV at {file_path}")
    
    logging.info('I am in save as csv function')
    return file_paths
    # Write CSV logic here

def save_as_excel(db_path, eventcodes_path, filename):
    # Ensure the filename has a .xlsx extension
    if not filename.lower().endswith('.xlsx'):
        filename = f"{filename}.xlsx"

    # Combine and convert the data into DataFrames
    final_dfs = combine_and_convert_df(db_path, eventcodes_path)
    
    if not final_dfs:
        print("No data fetched to save.")
        return
    
    shutil.rmtree(Config.TEMPORARY_FILES_DIR)
    save_dir = Config.TEMPORARY_FILES_DIR  # Temporary directory to save the file
    os.makedirs(save_dir, exist_ok=True)
    file_path = os.path.join(save_dir, filename)
    
    # Create an Excel writer object
    with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
        for table_name, df in final_dfs.items():
            df.to_excel(writer, sheet_name=table_name, index=False)
    
    # Open the saved workbook for formatting
    wb = load_workbook(file_path)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # Format first column: Convert Unix timestamp to dd/mm hh:mm:ss
        for row in ws.iter_rows(min_col=1, max_col=1, min_row=2):  # Skip header
            cell = row[0]
            if isinstance(cell.value, (int, float)):  # Check if it's a number
                dt = datetime.datetime.fromtimestamp(cell.value)
                cell.value = dt.strftime('%d/%m %H:%M:%S') + f'.{dt.microsecond // 1000:03d}'

        # Highlight headers
        header_fill = PatternFill(start_color="ffbd33", end_color="ffbd33", fill_type="solid")
        header_font = Font(bold=True)
        for cell in ws[1]:  # Assuming headers are in the first row
            cell.fill = header_fill
            cell.font = header_font

        # Center align all text
        center_alignment = Alignment(horizontal='center', vertical='center')
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = center_alignment

        def autofit_column_width(sheet):
            for col in sheet.columns:
                max_length = 0
                column = col[0].column_letter  # Get the column name (A, B, C, etc.)
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = (max_length + 2)  # Adding a little extra space
                sheet.column_dimensions[column].width = adjusted_width

        autofit_column_width(ws)

    wb.save(file_path)
    print(f"Data exported to Excel at {filename}")
    logging.info(file_path)
    
    return file_path